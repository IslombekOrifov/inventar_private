from typing import Type
from django.forms.models import BaseModelForm
from django.shortcuts import redirect, render, get_object_or_404, HttpResponse
from django.http import HttpResponseBadRequest, HttpResponseRedirect, Http404
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q, ProtectedError
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook
from django.contrib.auth.views import LoginView, PasswordChangeView
from django.contrib.auth.forms import PasswordChangeForm
from django.utils.decorators import method_decorator
from django.conf import settings

from accounts.models import CustomUser, Department
from index.models import Category, Model, Product
from .decorators import superuser_required
from .freshed_list import remove_from_list_item
from .forms import *


class CustomLoginView(LoginView):
    form_class = CustomAuthenticationForm
   

def test(request):
    return render(request, 'building.html')

def page_not_found_view(request, exception):
    return render(request, 'pages-404.html', status=404)

def custom_error_view(request, exception=None):
    return render(request, "pages-500.html", status=500)

# def custom_permission_denied_view(request, exception=None):
#     return render(request, "403.html", {})

# def custom_bad_request_view(request, exception=None):
#     return render(request, "errors/400.html", {})

@login_required
def Admin_index(request):
    if request.user.is_superuser:
        products = Product.objects.filter(group=request.user.groups.first()).filter(category_id__isnull=False)
        prod_for_count = Product.objects.filter(group=request.user.groups.first())
        product_count = len(prod_for_count)
        print(product_count)
        category_filt = Category.objects.filter(group=request.user.groups.first())
        categories = category_filt.annotate(count=Count('category'))
        query_count = products.filter(status=1).count()
        query_count_tw = products.filter(status=3).count()
        query_count_ir = products.filter(status=2).count()
        query_count_in = products.filter(status=0).count()

        page = request.GET.get('page', 1)
        paginator = Paginator(categories, 20)
        try:
            query = paginator.page(page)
        except PageNotAnInteger:
            query = paginator.page(1)
        except EmptyPage:
            query = paginator.page(paginator.num_pages)

        context = {
            'product_count': product_count,
            'query': query,
            'query_count': query_count, 
            'query_count_tw': query_count_tw,
            'query_count_ir': query_count_ir,
            'query_count_in': query_count_in,
            }
        return render(request, 'catalog_category.html', context=context)
    else:
        if request.user.status_member == CustomUser.UserStatus.WORKER:    
            return redirect('responsiblepage-myproducts')
        else:
            return redirect('responsiblepage-index')
    

@method_decorator(superuser_required, name='dispatch')
class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = Category
    form_class = CategoryCreateForm
    template_name = 'file/category-add.html'
    login_url = 'login'
    success_url = reverse_lazy('admin-categories')

    def form_valid(self, form):
        form.instance.author = self.request.user
        form.instance.group = self.request.user.groups.first()
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.request.user.groups.first()
        return context


@method_decorator(superuser_required, name='dispatch')
class ModelCreateView(LoginRequiredMixin, CreateView):
    model = Model
    form_class = ModelCreateForm
    template_name = 'file/model_add.html'
    login_url = 'login'
    success_url = reverse_lazy('admin-models')

    def form_valid(self, form):
        form.instance.author = self.request.user
        form.instance.group = self.request.user.groups.first()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.request.user.groups.first()
        return context
 

@login_required
def baseview(request, pk):
    q = Product.objects.filter(group=request.user.groups.first())
    prod_count = len(q)
    query_ = q.filter(category_id__id=pk).order_by('pk')
    get_cat = Category.objects.get(id=pk)
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'prod_count': prod_count,
        'query': query, 
        'get_cat': get_cat, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'index.html', context=context)

# PRODUCT BASED views

@login_required
def Unknown_Devices(request):
    q = Product.objects.filter(group=request.user.groups.first())
    query_ = q.filter(category_id=None)
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()
    if 'index/base/' in request:
        print(request)
    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)


    context = {
        'query': query, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'unknown_devices.html', context=context)


@login_required
@superuser_required 
def EquipmentCreateView(request, pk):
    products_count = Product.objects.filter(group=request.user.groups.first()).count()
    category = Category.objects.get(id=pk)
    inputs = request.POST.getlist('inputs')
    request.POST._mutable = True
    request.POST['description'] = ' | '.join(inputs)
    request.POST._mutable = False
    form = EquipmentCreateForm(request.user)
    if request.method == 'POST':
        form = EquipmentCreateForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data
            equipment = Product(
                group=request.user.groups.first(),
                name=cd['name'],
                schet=cd['schet'],
                category_id=category,
                room_number=cd['room_number'],
                inventar_number=cd['inventar_number'],
                model_id=cd['model_id'],
                responsible_id=cd['responsible_id'],
                images=cd['image'],
                seria_number=cd['seria_number'],
                year_of_manufacture=cd['year_of_manufacture'],
                unit_of_measurement=cd['unit_of_measurement'],
                description=cd['description'],
            )
            equipment.save()  
            red = category.id
            return redirect('base-view', pk=red)
        # else:
        #     return ("""<h1 style="color: "cyan"; text-align: "center";">Forma validatsiyadan o'tmadi!!!</h1>""")
    
    context = {
        'form': form, 
        'cat_id': category.id, 
        'products_count': products_count,
    }
    return render(request, 'file/product_add.html', context=context)


@login_required
def ProductDetailView(request, pk):
    eq = get_object_or_404(Product, pk=pk)
    descrip = eq.description.split(' | ')
    descript = remove_from_list_item(descrip, '|')
    product_qrcode = f"Inventar raqami: {eq.inventar_number} \n Javobgar shaxs: {'Nomalum' if eq.responsible is None else eq.responsible.email, eq.responsible.first_name, eq.responsible.last_name } \n Xona: {eq.room_number} \n "

    form = ProductDetailUpdateForm(request.user, request.POST or None, instance=eq)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('product-detail', pk=eq.pk)
    
    context = {
        'query': eq, 
        'product_qrcode': product_qrcode,
        'descript': descript, 
        'form': form,
    }
    return render(request, 'file/product-detail.html', context=context)


@login_required
@superuser_required 
def ProductUpdateView(request, pk):
    equipment = get_object_or_404(Product, pk=pk)
    if equipment.category_id is not None:
        red = equipment.category_id.id
    form = ProductUpdateForm(request.user, instance=equipment)
    if request.method == 'POST':
        if 'inputs' and 'status' in request.POST:
            inputs = request.POST.getlist('inputs')
            request.POST._mutable = True
            request.POST['status'] = request.POST['status']
            request.POST['description'] = equipment.description + ' | ' + ' | '.join(inputs)
            request.POST._mutable = False
        form = ProductUpdateForm(request.user, request.POST or None, request.FILES or None, instance=equipment)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        if equipment.category_id is not None:
            red = equipment.category_id.id
            return redirect('base-view', pk=red)
        else:
            return redirect('unknown-devices')
    
    context = {
        'form': form, 
        'equ': equipment,
    }

    if equipment.category_id is not None:
        context['red'] = red
        return render(request, 'file/product_update.html', context=context)
    else:
        return render(request, 'file/product_update.html', context=context)


@login_required
@superuser_required 
def ProductDeleteView(request, pk):
    query = Product.objects.get(pk=pk)
    if query.category_id is not None:
        red = query.category_id.id
    if request:
        query.delete()
        if query.category_id is not None:
            return redirect('base-view', pk=red)
        else:
            return redirect('unknown-devices')
    return render(request, 'admin/admin-base.html')


# SEARCh views

class SearchResultsView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'search-results.html'
    # paginate_by = 2

    def get_queryset(self):
        if self.request.GET.get('search') is not None:
            query = self.request.GET.get('search')
        else:
            query = ''
        object_list = Product.objects.filter(group=self.request.user.groups.first()).filter(
            Q(inventar_number__icontains=query) | Q(name__icontains=query)
        ).order_by('pk')
        return object_list
    
    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        query_count = self.object_list.filter(status=1).count()
        query_count_tw = self.object_list.filter(status=3).count()
        query_count_ir = self.object_list.filter(status=2).count()
        query_count_in = self.object_list.filter(status=0).count()
        context['query_count'] = query_count
        context['query_count_tw'] = query_count_tw
        context['query_count_ir'] = query_count_ir
        context['query_count_in'] = query_count_in
        return context

# CATEGORIES views
    
@login_required
def AdminCategories(request):
    query_ = Category.objects.filter(group=request.user.groups.first()).order_by('pk')
    categ_count = len(query_)
    query_p = Product.objects.filter(group=request.user.groups.first())
    query_count = query_p.filter(status=1).count()
    query_count_tw = query_p.filter(status=3).count()
    query_count_ir = query_p.filter(status=2).count()
    query_count_in = query_p.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 20)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'categ_count': categ_count, 
        'query': query, 
        'query_count': query_count, 
        'query_count_tw': query_count_tw, 
        'query_count_ir': query_count_ir, 
        'query_count_in': query_count_in,
    }

    return render(request, 'card.html', context=context)


@login_required
@superuser_required 
def AdminCategoryEdit(request, pk):
    query = get_object_or_404(Category, pk=pk)
    form = CategoryEditForm(request.POST or None, request.FILES or None, instance=query)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('admin-categories')

    context = {
        'form': form,
    }
    return render(request, 'file/category_edit.html', context=context)


@login_required
@superuser_required 
def AdminCategoryDelete(request, pk):
    query = Category.objects.get(pk=pk)
    if request:
        query.delete()
        return redirect('admin-categories')
    return render(request, 'card.html')


# UNIVERSAL EXCEL UPLODER VIEW
@login_required
@superuser_required 
def UploadExcelFile(request):
    if request.method == "POST":
        user_gp = request.user.groups.first()
        form = UploadExcelFileForm(request.POST, request.FILES)
        if form.is_valid():
            if form.cleaned_data['excel_file_cat'] is not None:
                wb = load_workbook(filename=request.FILES['excel_file_cat'].file)
                worksheet = wb["Sheet1"]
                for i in worksheet.iter_rows(min_row=2):
                    category_count = Category.objects.filter(group=user_gp)
                    if user_gp.max_limit_category == len(category_count):
                        error = _("Siz boshqa kategoriya qo'sha olmaysiz!")
                        return redirect("category-create", {'error': error})
                        break
                    else:
                        result = Category(
                            group = user_gp,
                            name_uz = i[0].value,
                            name_en = i[1].value,
                            name_ru = i[2].value,
                        )
                        result.save()
                        return redirect("admin-categories")
            elif form.cleaned_data['excel_file_mod'] is not None:
                wb = load_workbook(filename=request.FILES['excel_file_mod'].file)
                worksheet = wb["Sheet1"]
                for i in worksheet.iter_rows(min_row=2):
                    model_count = Model.objects.filter(group=user_gp)
                    if user_gp.max_limit_model > len(model_count):
                        result = Model(
                            group = user_gp,
                            name = i[0].value,
                            description = i[1].value,
                        )
                        result.save()
                return redirect("admin-models")
            elif form.cleaned_data['excel_file_depart'] is not None:
                wb = load_workbook(filename=request.FILES['excel_file_depart'].file)
                worksheet = wb["Sheet1"]
                for i in worksheet.iter_rows(min_row=2):
                    model_count = Department.objects.filter(group=user_gp)
                    if user_gp.max_limit_department > len(model_count):
                        result = Department(
                            group = user_gp,
                            name = i[0].value,
                        )
                        result.save()
                return redirect("department-list")
            elif form.cleaned_data['excel_file_pro'] is not None:
                try:
                    wb = load_workbook(filename=request.FILES['excel_file_pro'].file)
                    worksheet = wb["Sheet1"]
                    for i in worksheet.iter_rows(min_row=2):
                        product_count = Product.objects.filter(group=user_gp)
                        if user_gp.max_limit_product > len(product_count):
                            result = Product(
                                group = request.user.groups.first(),
                                name = i[0].value,
                                schet = i[1].value,
                                inventar_number = i[2].value,
                                seria_number = i[3].value,
                                description = i[4].value,
                                year_of_manufacture = i[5].value,
                                unit_of_measurement = i[6].value,
                                room_number = i[7].value,
                                product_count = i[8].value,
                            )
                            result.save()
                    return redirect("unknown-devices")
                except Exception as e:
                    error = _("Faylda bazada mavjud inventar raqamga ega jihoz mavjud!")
                    return render(request, 'file/product_add_excel.html', {'error': error, 'e': e})
            else:
                return HttpResponseBadRequest()
        else:
            return HttpResponseBadRequest()
    else:
        return render(request, 'unknown_devices.html', {})

def product_create_excel(request):
    return render(request, 'file/product_add_excel.html')
    

# MODELS views
@login_required
def AdminModels(request):
    query_ = Model.objects.filter(group=request.user.groups.first()).order_by('pk')
    query_p = Product.objects.filter(group=request.user.groups.first())
    query_count = query_p.filter(status=1).count()
    query_count_tw = query_p.filter(status=3).count()
    query_count_ir = query_p.filter(status=2).count()
    query_count_in = query_p.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 20)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'query': query, 
        'query_count': query_count, 
        'query_count_tw': query_count_tw, 
        'query_count_ir': query_count_ir, 
        'query_count_in': query_count_in,
    }

    return render(request, 'models.html', context=context)


@login_required
@superuser_required 
def AdminModelEdit(request, pk):
    query = get_object_or_404(Model, pk=pk)
    form = ModelEditForm(request.POST or None, request.FILES or None, instance=query)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('admin-models')

    context = {
        'form': form,
    }
    return render(request, 'file/models_edit.html', context=context)


@login_required
@superuser_required 
def AdminModelDelete(request, pk):
    query = Model.objects.get(pk=pk, group=request.user.groups.first())
    if request.user.is_superuser:
        query.delete()
    return redirect('admin-models')


@login_required
@superuser_required 
def AdminResponsibleDelete(request, pk):
    query = CustomUser.objects.get(
        ~Q(status=CustomUser.UserStatus.VIEWER),
        pk=pk, groups=request.user.groups.first(), 
        is_superuser=False, 
    )
    if query.exists():
        query.delete()
        return redirect('pages:responsible')
    return render(request, 'responsible_person.html')


# ROOM views
@login_required
def AdminRooms(request):
    user = request.user.groups.first()
    queryset = Product.objects.filter(group=user).order_by('room_number')
    room_list = []
    for u in queryset:
        if u.room_number not in room_list:
            room_list.append(u.room_number)
    context = {
        'user': user,
        'queryset': queryset,
        'room_list': room_list,
    }
    return render(request, 'rooms_catalog.html', context=context)


@login_required
def AdminRoomFloor(request, pk):
    
    context = {
        'pka': pk,
    }
    return render(request, 'rooms_catalog_floor.html', context=context)


@login_required
def AdminRoomDetail(request, slug):
    prod = request.user.groups.first()
    if prod.ui == True:
        query = Product.objects.filter(
            Q(room_number__icontains=slug)
        ).order_by('pk')
    else:
        query = Product.objects.filter(room_number=slug).order_by('pk')

    query_count = query.filter(status=1).count()
    query_count_tw = query.filter(status=3).count()
    query_count_ir = query.filter(status=2).count()
    query_count_in = query.filter(status=0).count()

    context = {
        'query': query,
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'file/room_devices.html', context=context)


@login_required
def profile_edit(request):
    pass_form = PasswordChangeForm(user=request.user)
    if request.method == 'POST':
        profile_form = ProfileEditForm(instance=request.user, data=request.POST)
        if profile_form.is_valid():
            user = profile_form.save(commit=False)
            user.username = profile_form.cleaned_data['email']
            user.save()
    else:
        profile_form = ProfileEditForm(instance=request.user)
    context = {
        'profile_form': profile_form,
        'pass_form': pass_form,  
        'password_change_done': False                  
    }
    return render(request, 'file/profile-update.html', context)

    

@login_required
def password_change(request):
    profile_form = ProfileEditForm(instance=request.user)
    password_change_done = False
    if request.method == 'POST':
        pass_form = PasswordChangeForm(user=request.user, data=request.POST)
        if pass_form.is_valid():
            user = CustomUser.objects.get(id=request.user.id)
            if user.check_password(pass_form.cleaned_data['old_password']):
                user.set_password(pass_form.cleaned_data['new_password2'])
                user.save()
                update_session_auth_hash(request, user)
                password_change_done = True
    else:
        pass_form = PasswordChangeForm(user=request.user)
    context = {
        'profile_form': profile_form,
        'pass_form': pass_form,  
        'password_change_done': password_change_done,              
    }
    return render(request, 'file/profile-update.html', context)
    
    

@login_required
def responsible_baseview(request):
    if not request.user.is_superuser and request.user.status_member == CustomUser.UserStatus.LEADER:
        q = Product.objects.filter(
            Q(responsible__status_member=CustomUser.UserStatus.WORKER) | 
            Q(responsible__status_member=CustomUser.UserStatus.LEADER), 
            group=request.user.groups.first(),
            responsible__department=request.user.department
        )
    else:
        q = Product.objects.filter(group=request.user.groups.first())
    prod_count = len(q)
    query_ = q.order_by('pk')
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(q, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'prod_count': prod_count,
        'query': query, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'index.html', context=context)


@login_required
def responsible_myproduct(request):
    q = Product.objects.filter(
        responsible__pk=request.user.id, 
        group=request.user.groups.first()             
    )
    prod_count = len(q)
    query_ = q.order_by('pk')
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(q, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'prod_count': prod_count,
        'query': query, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'index.html', context=context)


@login_required
def responsible_application(request):
    q = Product.objects.filter(responsible_id=request.user.id)
    prod_count = len(q)
    # query_ = q.filter(category_id__id=pk).order_by('pk')
    # get_cat = Category.objects.get(id=pk)
    # query_count = query_.filter(status=1).count()
    # query_count_tw = query_.filter(status=3).count()
    # query_count_ir = query_.filter(status=2).count()
    # query_count_in = query_.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(q, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'prod_count': prod_count,
        'query': query, 
        'get_cat': 0, 
        'query_count': 0,
        'query_count_tw': 0,
        'query_count_ir': 0,
        'query_count_in': 0,
        }
    return render(request, 'responsiblepage-application.html', context=context)



@login_required
def responsible_application_create(request):
    products_count = Product.objects.filter(group=request.user.groups.first()).count()
    category = Category.objects.all().first()
    inputs = request.POST.getlist('inputs')
    request.POST._mutable = True
    request.POST['description'] = ' | '.join(inputs)
    request.POST._mutable = False
    form = EquipmentCreateForm(request.user)
    if request.method == 'POST':
        form = EquipmentCreateForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data
            equipment = Product(
                group=request.user.groups.first(),
                name=cd['name'],
                schet=cd['schet'],
                category_id=category,
                room_number=cd['room_number'],
                inventar_number=cd['inventar_number'],
                model_id=cd['model_id'],
                responsible_id=cd['responsible_id'],
                images=cd['image'],
                seria_number=cd['seria_number'],
                year_of_manufacture=cd['year_of_manufacture'],
                unit_of_measurement=cd['unit_of_measurement'],
                description=cd['description'],
            )
            equipment.save()  
            red = category.id
            return redirect('base-view', pk=red)
        # else:
        #     return ("""<h1 style="color: "cyan"; text-align: "center";">Forma validatsiyadan o'tmadi!!!</h1>""")
    
    context = {
        'form': form, 
        'cat_id': category.id, 
        'products_count': products_count,
    }
    return render(request, 'file/product_add.html', context=context)


# Department and RESPONSIBLE PERSONS views
@login_required
def department_list(request):
    if request.user.is_superuser or request.user.status_member in ['dr', 'dp']:
        depart_for_count = Department.objects.filter(group=request.user.groups.first())
        depart_count = len(depart_for_count)
        departments = depart_for_count.annotate(count=Count('users'))

        page = request.GET.get('page', 1)
        paginator = Paginator(departments, 20)
        try:
            query = paginator.page(page)
        except PageNotAnInteger:
            query = paginator.page(1)
        except EmptyPage:
            query = paginator.page(paginator.num_pages)

        context = {
            'depart_count': depart_count,
            'query': query,
            }
        return render(request, 'catalog_department.html', context=context)
    else:
        return redirect('responsiblepage-index')


@method_decorator(superuser_required, name='dispatch')
class DepartmentCreateView(LoginRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentCreateForm
    template_name = 'file/department-add.html'
    login_url = 'login'
    success_url = reverse_lazy('department-list')

    def form_valid(self, form):
        form.instance.group = self.request.user.groups.first()
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.request.user.groups.first()
        return context
    

@login_required
@superuser_required 
def department_edit(request, pk):
    query = Department.objects.get(pk=pk, group=request.user.groups.first())
    form = DepartmentCreateForm(instance=query)
    if request.method == 'POST':
        form = DepartmentCreateForm(request.POST, instance=query)
        if form.is_valid():
            print(form.cleaned_data)
            depart = form.save(commit=False)
            depart.group=request.user.groups.first()
            depart.save()
            return redirect('department-list')
    context = {
        'form': form,
    }
    return render(request, 'file/department-edit.html', context=context)


@login_required
@superuser_required 
def AdminDepartmentDelete(request, pk):
    query = Department.objects.get(
        pk=pk, group=request.user.groups.first(), 
    )
    if query.exists():
        try:
            query.delete()
            return redirect('department-list')
        except ProtectedError:
            return redirect('department-list')
    return redirect('department-list')


@login_required
def AdminResponsibles(request, pk=None):
    if not request.user.is_superuser and request.user.status_member in ['lr', 'dr', 'dp']:
        if pk:
            query_ = CustomUser.objects.filter(
                groups=request.user.groups.first(), department__pk=pk,
                is_superuser=False
            ).order_by('pk')
        else:
            query_ = CustomUser.objects.filter(
                groups=request.user.groups.first(), 
                department__pk=request.user.department.id,
                status_member=CustomUser.UserStatus.WORKER,
                is_superuser=False
            ).order_by('pk')
    elif request.user.is_superuser:
        query_ = CustomUser.objects.filter(
            groups=request.user.groups.first(), department__pk=pk,
            is_superuser=False
        ).order_by('pk')
    else:
        raise Http404()
    
    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)
    print(query)
    context = {
        'query': query, 
        'depart_id': pk,
    }
    return render(request, 'responsible_person.html', context=context)


@method_decorator(superuser_required, name='dispatch')
class ResponsibleCreateView(LoginRequiredMixin, CreateView):
    model = CustomUser
    form_class = ResponsibleCreateForm
    template_name = 'file/responsible_add.html'
    login_url = 'login'
    success_url = reverse_lazy('admin-responsibles')
    
    def form_valid(self, form):
        if 'pk' in self.kwargs:
            depart_id = self.kwargs['pk']
            if depart_id:
                self.success_url = reverse_lazy('admin-responsibles', args=[depart_id])
                department = get_object_or_404(Department, pk=depart_id)
                form.instance.department = department
        form.instance.author = self.request.user
        return super().form_valid(form)

    def get_success_url(self) -> str:
        department = Department.objects.get(users__id=self.object.pk)
        print(department)
        self.success_url = reverse_lazy('admin-responsibles', args=[department.id])
        self.object.groups.add(self.request.user.groups.first())
        return super().get_success_url()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['depart_id'] = self.kwargs.get('pk', False)
        group = self.request.user.groups.first()
        return context
    

@login_required
@superuser_required 
def AdminResponsibleEdit(request, pk):
    query = get_object_or_404(CustomUser, pk=pk, is_superuser=False, groups=request.user.groups.first())
    form = ResponsibleEditForm(request.POST or None, instance=query)
    pass_form = PasswordChangeForm(user=query)
    if request.method == 'POST' and form.is_valid():
        user = form.save(commit=False)
        user.username = form.cleaned_data['email']
        user.save()
        return redirect('admin-responsibles', user.department.id)
    context = {
        'user': query,
        'form': form,
        'pass_form': pass_form,
        'depart_id': query.department.id,
    }
    return render(request, 'file/responsible-edit.html', context=context)


@login_required
@superuser_required 
def responsible_password_change(request, pk):
    user = get_object_or_404(CustomUser, pk=pk, is_superuser=False, groups=request.user.groups.first())
    profile_form = ResponsibleEditForm(instance=user)
    if request.user.is_superuser:
        password_change_done = False
        if request.method == 'POST':
            pass_form = PasswordChangeForm(user=user, data=request.POST)
            if pass_form.is_valid():
                if user.check_password(pass_form.cleaned_data['old_password']):
                    user.set_password(pass_form.cleaned_data['new_password2'])
                    user.save()
                    password_change_done = True
        else:
            pass_form = PasswordChangeForm(user=user)
        context = {
            'user': user,
            'form': profile_form,
            'pass_form': pass_form,  
            'password_change_done': password_change_done, 
            'depart_id': user.department.id,             
        }
        return render(request, 'file/responsible-edit.html', context)
    else:
        raise Http404()


@login_required
def AdminResponsibleProducts(request, pk):
    if request.user.is_superuser or request.user.status_member in ['lr', 'dp', 'dr']:
        q = Product.objects.filter(group=request.user.groups.first())
        prod_count = len(q)
        query_ = q.filter(responsible__id=pk).order_by('pk')
        get_cat = CustomUser.objects.get(id=pk, groups=request.user.groups.first())
        query_count = query_.filter(status=1).count()
        query_count_tw = query_.filter(status=3).count()
        query_count_ir = query_.filter(status=2).count()
        query_count_in = query_.filter(status=0).count()

        page = request.GET.get('page', 1)
        paginator = Paginator(query_, 50)
        try:
            query = paginator.page(page)
        except PageNotAnInteger:
            query = paginator.page(1)
        except EmptyPage:
            query = paginator.page(paginator.num_pages)

        context = {
            'prod_count': prod_count,
            'query': query, 
            'get_cat': get_cat, 
            'query_count': query_count,
            'query_count_tw': query_count_tw,
            'query_count_ir': query_count_ir,
            'query_count_in': query_count_in,
            }
        return render(request, 'index-responsible.html', context=context)
    else: 
        raise Http404()